"""Articles API routes for managing article operations.

This module provides REST API endpoints for managing articles including:
- Listing articles with job-based filtering
- Article search and metadata display
- Article export functionality in multiple formats
- Article statistics and analytics

All endpoints support proper pagination, filtering, and error handling
with correlation IDs for request tracking.

Example:
    Using the articles API:

    ```bash
    # List articles by job ID
    GET /api/v1/articles?job_id=123e4567-e89b-12d3-a456-426614174000

    # Search articles
    GET /api/v1/articles?search=python&min_relevance_score=0.5

    # Export articles
    POST /api/v1/articles/export
    {
        "job_id": "123e4567-e89b-12d3-a456-426614174000",
        "format": "csv"
    }
    ```
"""

import logging
import io
import csv
import json
from typing import Optional, List, Dict, Any
from uuid import UUID
from datetime import datetime, timezone

from fastapi import APIRouter, HTTPException, Depends, Request, Query, Response
from fastapi.responses import StreamingResponse, JSONResponse
import structlog

from src.api.schemas.article import (
    ArticleResponse,
    ArticleListResponse,
    ArticleSearchRequest,
    ArticleExportRequest,
    ArticleExportResponse,
    ArticleStatsResponse,
    ErrorResponse
)
from src.database.repositories.article_repo import ArticleRepository
from src.database.repositories.job_repo import CrawlJobRepository
from src.database.repositories.category_repo import CategoryRepository
from src.shared.exceptions import (
    BaseAppException,
    DatabaseConnectionError,
    ValidationError,
    JobNotFoundException,
    CategoryNotFoundError
)

logger = structlog.get_logger(__name__)

router = APIRouter(prefix="/api/v1", tags=["articles"])


def get_article_repository() -> ArticleRepository:
    """Dependency to get ArticleRepository instance."""
    return ArticleRepository()


def get_job_repository() -> CrawlJobRepository:
    """Dependency to get CrawlJobRepository instance."""
    return CrawlJobRepository()


def get_category_repository() -> CategoryRepository:
    """Dependency to get CategoryRepository instance."""
    return CategoryRepository()


@router.get("/articles", response_model=ArticleListResponse)
async def list_articles(
    request: Request,
    job_id: Optional[UUID] = Query(None, description="Filter by job ID"),
    category_id: Optional[UUID] = Query(None, description="Filter by category ID"),
    search: Optional[str] = Query(None, description="Search in title and content"),
    keywords: Optional[str] = Query(None, description="Filter by keywords (comma-separated)"),
    min_relevance_score: Optional[float] = Query(None, ge=0.0, le=1.0, description="Minimum relevance score"),
    from_date: Optional[datetime] = Query(None, description="Filter from date (ISO format)"),
    to_date: Optional[datetime] = Query(None, description="Filter to date (ISO format)"),
    page: int = Query(1, ge=1, description="Page number"),
    size: int = Query(20, ge=1, le=100, description="Page size"),
    article_repo: ArticleRepository = Depends(get_article_repository),
    job_repo: CrawlJobRepository = Depends(get_job_repository),
    category_repo: CategoryRepository = Depends(get_category_repository)
) -> ArticleListResponse:
    """List articles with filtering and pagination.

    This endpoint returns a paginated list of articles with optional filtering
    by job ID, category, search terms, and other criteria. Supports full-text
    search across title and content fields.

    Args:
        request: FastAPI request object for correlation ID
        job_id: Optional job UUID filter
        category_id: Optional category UUID filter
        search: Optional search query for title/content
        keywords: Optional keywords filter (comma-separated)
        min_relevance_score: Optional minimum relevance score
        from_date: Optional start date filter
        to_date: Optional end date filter
        page: Page number (1-based)
        size: Number of articles per page
        article_repo: Article repository dependency
        job_repo: Job repository dependency
        category_repo: Category repository dependency

    Returns:
        Paginated list of articles with metadata

    Raises:
        HTTPException: If filtering or retrieval fails
    """
    correlation_id = getattr(request.state, 'correlation_id', 'unknown')

    logger.info(
        "Listing articles",
        correlation_id=correlation_id,
        job_id=str(job_id) if job_id else None,
        category_id=str(category_id) if category_id else None,
        search_query=search,
        page=page,
        size=size
    )

    try:
        # Validate job_id if provided
        if job_id:
            job = await job_repo.get_by_id(job_id)
            if not job:
                raise HTTPException(
                    status_code=404,
                    detail=f"Job with ID {job_id} not found"
                )

        # Validate category_id if provided
        if category_id:
            category = await category_repo.get_by_id(category_id)
            if not category:
                raise HTTPException(
                    status_code=404,
                    detail=f"Category with ID {category_id} not found"
                )

        # Parse keywords if provided
        keywords_list = None
        if keywords:
            keywords_list = [k.strip() for k in keywords.split(',') if k.strip()]

        # Build search filters
        filters = {
            'job_id': job_id,
            'category_id': category_id,
            'search_query': search,
            'keywords': keywords_list,
            'min_relevance_score': min_relevance_score,
            'from_date': from_date,
            'to_date': to_date
        }

        # Remove None values
        filters = {k: v for k, v in filters.items() if v is not None}

        # Get articles with pagination
        articles, total = await article_repo.get_articles_paginated(
            filters=filters,
            page=page,
            size=size
        )

        # Convert to response format
        article_responses = []
        for article in articles:
            # Build categories list from ArticleCategory relationships
            categories_info = []
            primary_cat_id = None

            # Get category_id from crawl_job if available
            if article.crawl_job_id and job_id:
                job = await job_repo.get_by_id(article.crawl_job_id)
                if job:
                    primary_cat_id = str(job.category_id) if job.category_id else None

            # Process article categories
            for article_cat in (article.categories or []):
                # Load category details
                cat = await category_repo.get_by_id(article_cat.category_id)
                if cat:
                    categories_info.append({
                        'id': str(article_cat.category_id),
                        'name': cat.name,
                        'relevance_score': float(article_cat.relevance_score) if article_cat.relevance_score else 1.0
                    })

                    # Set primary category if not already set
                    if not primary_cat_id:
                        primary_cat_id = str(article_cat.category_id)

            article_responses.append(
                ArticleResponse(
                    id=article.id,
                    title=article.title,
                    content=article.content,
                    author=article.author,
                    publish_date=article.publish_date,
                    source_url=article.source_url,
                    image_url=article.image_url,
                    url_hash=article.url_hash,
                    content_hash=article.content_hash,
                    last_seen=article.last_seen,
                    crawl_job_id=article.crawl_job_id,
                    keywords_matched=article.keywords_matched or [],
                    relevance_score=article.relevance_score,
                    categories=categories_info if categories_info else None,
                    primary_category_id=primary_cat_id,
                    created_at=article.created_at,
                    updated_at=article.updated_at
                )
            )

        # Calculate total pages
        total_pages = (total + size - 1) // size

        logger.info(
            "Articles retrieved successfully",
            correlation_id=correlation_id,
            total_returned=len(article_responses),
            total_count=total,
            page=page,
            total_pages=total_pages
        )

        return ArticleListResponse(
            articles=article_responses,
            total=total,
            page=page,
            size=size,
            pages=total_pages
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(
            "Failed to list articles",
            correlation_id=correlation_id,
            error=str(e),
            error_type=type(e).__name__
        )
        raise HTTPException(
            status_code=500,
            detail=f"Failed to retrieve articles: {str(e)}"
        )


@router.get("/articles/{article_id}", response_model=ArticleResponse)
async def get_article(
    article_id: UUID,
    request: Request,
    article_repo: ArticleRepository = Depends(get_article_repository)
) -> ArticleResponse:
    """Get detailed information for a specific article.

    Args:
        article_id: UUID of the article to retrieve
        request: FastAPI request object for correlation ID
        article_repo: Article repository dependency

    Returns:
        Article details with metadata

    Raises:
        HTTPException: If article not found or retrieval fails
    """
    correlation_id = getattr(request.state, 'correlation_id', 'unknown')

    logger.info(
        "Getting article details",
        correlation_id=correlation_id,
        article_id=str(article_id)
    )

    try:
        article = await article_repo.get_by_id(article_id)
        if not article:
            logger.warning(
                "Article not found",
                correlation_id=correlation_id,
                article_id=str(article_id)
            )
            raise HTTPException(
                status_code=404,
                detail=f"Article with ID {article_id} not found"
            )

        logger.info(
            "Article retrieved successfully",
            correlation_id=correlation_id,
            article_id=str(article_id)
        )

        return ArticleResponse(
            id=article.id,
            title=article.title,
            content=article.content,
            author=article.author,
            publish_date=article.publish_date,
            source_url=article.source_url,
            image_url=article.image_url,
            url_hash=article.url_hash,
            content_hash=article.content_hash,
            last_seen=article.last_seen,
            crawl_job_id=article.crawl_job_id,
            keywords_matched=article.keywords_matched or [],
            relevance_score=article.relevance_score,
            created_at=article.created_at,
            updated_at=article.updated_at
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(
            "Failed to get article",
            correlation_id=correlation_id,
            article_id=str(article_id),
            error=str(e),
            error_type=type(e).__name__
        )
        raise HTTPException(
            status_code=500,
            detail=f"Failed to retrieve article: {str(e)}"
        )


@router.post("/articles/export")
async def export_articles(
    export_request: ArticleExportRequest,
    request: Request,
    article_repo: ArticleRepository = Depends(get_article_repository),
    job_repo: CrawlJobRepository = Depends(get_job_repository),
    category_repo: CategoryRepository = Depends(get_category_repository)
) -> StreamingResponse:
    """Export articles in various formats (JSON, CSV, Excel).

    This endpoint exports articles based on filtering criteria and returns
    the file as a streaming response with proper UTF-8 encoding for Vietnamese
    characters support.

    Args:
        export_request: Export configuration and filters
        request: FastAPI request object for correlation ID
        article_repo: Article repository dependency
        job_repo: Job repository dependency
        category_repo: Category repository dependency

    Returns:
        Streaming response with exported file

    Raises:
        HTTPException: If export fails or invalid parameters
    """
    correlation_id = getattr(request.state, 'correlation_id', 'unknown')

    logger.info(
        "Exporting articles",
        correlation_id=correlation_id,
        job_id=str(export_request.job_id) if export_request.job_id else None,
        category_id=str(export_request.category_id) if export_request.category_id else None,
        format=export_request.format
    )

    try:
        # Validate job_id if provided
        if export_request.job_id:
            job = await job_repo.get_by_id(export_request.job_id)
            if not job:
                raise HTTPException(
                    status_code=404,
                    detail=f"Job with ID {export_request.job_id} not found"
                )

        # Validate category_id if provided
        if export_request.category_id:
            category = await category_repo.get_by_id(export_request.category_id)
            if not category:
                raise HTTPException(
                    status_code=404,
                    detail=f"Category with ID {export_request.category_id} not found"
                )

        # Build filters for export
        filters = {}
        if export_request.job_id:
            filters['job_id'] = export_request.job_id
        if export_request.category_id:
            filters['category_id'] = export_request.category_id
        if export_request.search_filters:
            # Add additional search filters if provided
            search_filters = export_request.search_filters.dict(exclude_unset=True)
            filters.update(search_filters)

        # Get all articles matching filters (no pagination for export)
        articles, total = await article_repo.get_articles_paginated(
            filters=filters,
            page=1,
            size=10000  # Large number to get all articles
        )

        if not articles:
            raise HTTPException(
                status_code=404,
                detail="No articles found matching the specified criteria"
            )

        # Determine fields to export
        fields = export_request.fields or [
            'id', 'title', 'author', 'publish_date', 'source_url',
            'keywords_matched', 'relevance_score', 'created_at'
        ]

        # Generate filename
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"articles_export_{timestamp}.{export_request.format}"

        # Export based on format
        if export_request.format == "json":
            content = export_to_json(articles, fields)
            media_type = "application/json"
        elif export_request.format == "csv":
            content = export_to_csv(articles, fields)
            media_type = "text/csv"
        elif export_request.format == "xlsx":
            content = export_to_excel(articles, fields)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported export format: {export_request.format}"
            )

        logger.info(
            "Articles exported successfully",
            correlation_id=correlation_id,
            total_articles=len(articles),
            format=export_request.format,
            filename=filename
        )

        # Return streaming response with proper headers
        headers = {
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Type": media_type,
            "X-Total-Articles": str(len(articles)),
            "X-Correlation-ID": correlation_id
        }

        return StreamingResponse(
            io.BytesIO(content),
            media_type=media_type,
            headers=headers
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(
            "Failed to export articles",
            correlation_id=correlation_id,
            error=str(e),
            error_type=type(e).__name__
        )
        raise HTTPException(
            status_code=500,
            detail=f"Failed to export articles: {str(e)}"
        )


def export_to_json(articles: List, fields: List[str]) -> bytes:
    """Export articles to JSON format with UTF-8 encoding."""
    article_data = []
    for article in articles:
        article_dict = {}
        for field in fields:
            value = getattr(article, field, None)
            if isinstance(value, datetime):
                value = value.isoformat()
            elif isinstance(value, UUID):
                value = str(value)
            article_dict[field] = value
        article_data.append(article_dict)

    json_content = json.dumps(article_data, ensure_ascii=False, indent=2)
    return json_content.encode('utf-8')


def export_to_csv(articles: List, fields: List[str]) -> bytes:
    """Export articles to CSV format with UTF-8 encoding."""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fields)
    writer.writeheader()

    for article in articles:
        row = {}
        for field in fields:
            value = getattr(article, field, None)
            if isinstance(value, datetime):
                value = value.isoformat()
            elif isinstance(value, UUID):
                value = str(value)
            elif isinstance(value, list):
                value = ', '.join(str(item) for item in value) if value else ''
            row[field] = value
        writer.writerow(row)

    csv_content = output.getvalue()
    output.close()
    return csv_content.encode('utf-8')


def export_to_excel(articles: List, fields: List[str]) -> bytes:
    """Export articles to Excel format with UTF-8 encoding."""
    try:
        import openpyxl
        from openpyxl.workbook import Workbook
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel export requires openpyxl package"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Articles Export"

    # Write headers
    for col, field in enumerate(fields, 1):
        ws.cell(row=1, column=col, value=field)

    # Write data
    for row, article in enumerate(articles, 2):
        for col, field in enumerate(fields, 1):
            value = getattr(article, field, None)
            if isinstance(value, datetime):
                value = value.isoformat()
            elif isinstance(value, UUID):
                value = str(value)
            elif isinstance(value, list):
                value = ', '.join(str(item) for item in value) if value else ''
            ws.cell(row=row, column=col, value=value)

    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


@router.get("/articles/stats", response_model=ArticleStatsResponse)
async def get_article_stats(
    request: Request,
    article_repo: ArticleRepository = Depends(get_article_repository)
) -> ArticleStatsResponse:
    """Get article statistics and analytics.

    Args:
        request: FastAPI request object for correlation ID
        article_repo: Article repository dependency

    Returns:
        Article statistics including counts and averages

    Raises:
        HTTPException: If stats retrieval fails
    """
    correlation_id = getattr(request.state, 'correlation_id', 'unknown')

    logger.info(
        "Getting article statistics",
        correlation_id=correlation_id
    )

    try:
        stats = await article_repo.get_article_statistics()

        logger.info(
            "Article statistics retrieved successfully",
            correlation_id=correlation_id,
            total_articles=stats.get('total_articles', 0)
        )

        return ArticleStatsResponse(**stats)

    except Exception as e:
        logger.error(
            "Failed to get article statistics",
            correlation_id=correlation_id,
            error=str(e),
            error_type=type(e).__name__
        )
        raise HTTPException(
            status_code=500,
            detail=f"Failed to retrieve article statistics: {str(e)}"
        )